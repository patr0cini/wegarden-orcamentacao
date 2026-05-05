import base64, io, os, json, secrets
from flask import Flask, request, send_file, session, redirect, send_from_directory, jsonify
import openpyxl
import urllib.request, urllib.parse

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, 'static') if os.path.exists(os.path.join(BASE_DIR, 'static', 'index.html')) else BASE_DIR

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'wegarden-secret-2026')
APP_PASSWORD        = os.environ.get('APP_PASSWORD', 'wegarden2026')
AZURE_CLIENT_ID     = os.environ.get('AZURE_CLIENT_ID', '')
AZURE_CLIENT_SECRET = os.environ.get('AZURE_CLIENT_SECRET', '')
AZURE_TENANT_ID     = os.environ.get('AZURE_TENANT_ID', '')
APP_BASE_URL        = os.environ.get('APP_BASE_URL', 'https://wegarden-orcamentacao.onrender.com')
SP_SCOPES           = 'https://graph.microsoft.com/Sites.ReadWrite.All offline_access openid'

# In-memory token store: { short_token -> access_token }
# Short tokens are 1-time use, expire after first read
_TOKEN_STORE = {}

def authed(): return session.get('auth') is True

def login_page(error=''):
    err_html = f'<div class="err">{error}</div>' if error else ''
    return f'''<!DOCTYPE html>
<html lang="pt"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>We Garden</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Instrument+Sans:wght@400;500;600&display=swap');
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Instrument Sans',sans-serif;background:#f0ede6;display:flex;align-items:center;justify-content:center;min-height:100vh}}
.card{{background:#fff;border-radius:16px;padding:2.5rem 2rem;width:100%;max-width:360px;box-shadow:0 4px 24px rgba(0,0,0,.08)}}
.logo{{display:flex;align-items:center;gap:10px;margin-bottom:2rem}}
.dot{{width:32px;height:32px;background:#3d7a52;border-radius:9px;display:flex;align-items:center;justify-content:center;font-size:17px}}
.lt{{font-size:17px;font-weight:600;color:#1e3a28}}
h1{{font-size:15px;font-weight:500;color:#4a5249;margin-bottom:1.25rem}}
input{{width:100%;padding:10px 12px;border:1px solid #d0ccc4;border-radius:8px;font-family:inherit;font-size:14px;background:#faf9f6;outline:none;margin-bottom:1rem}}
input:focus{{border-color:#3d7a52;box-shadow:0 0 0 3px #eaf2ec}}
button{{width:100%;padding:10px;background:#1e3a28;color:#fff;border:none;border-radius:8px;font-family:inherit;font-size:14px;font-weight:600;cursor:pointer}}
button:hover{{background:#2d5a3d}}
.err{{color:#8b2020;font-size:13px;margin-bottom:.75rem;background:#fef0ee;padding:8px 12px;border-radius:6px}}
</style></head><body>
<div class="card">
  <div class="logo"><div class="dot">&#127807;</div><div class="lt">We Garden</div></div>
  <h1>Introduz a password para aceder</h1>
  {err_html}
  <form method="post">
    <input type="password" name="password" placeholder="Password" autofocus autocomplete="current-password">
    <button type="submit">Entrar &rarr;</button>
  </form>
</div></body></html>'''

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        if request.form.get('password') == APP_PASSWORD:
            session['auth'] = True
            return redirect('/')
        return login_page('Password incorrecta.')
    return login_page()

@app.route('/logout')
def logout():
    session.clear(); return redirect('/login')

@app.route('/')
def index():
    if not authed(): return redirect('/login')
    return send_from_directory(STATIC_DIR, 'index.html')

# ── SHAREPOINT OAUTH (popup-friendly) ────────────────────────────
@app.route('/auth/sharepoint')
def auth_sharepoint():
    # This route is opened in a popup — no auth check needed
    if not AZURE_CLIENT_ID:
        return 'Azure AD não configurado.', 500
    state = secrets.token_urlsafe(16)
    session['oauth_state'] = state
    params = urllib.parse.urlencode({
        'client_id':     AZURE_CLIENT_ID,
        'response_type': 'code',
        'redirect_uri':  APP_BASE_URL + '/auth/callback',
        'scope':         SP_SCOPES,
        'state':         state,
        'prompt':        'select_account'
    })
    return redirect(f'https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/authorize?{params}')

@app.route('/auth/callback')
def auth_callback():
    error = request.args.get('error')
    if error:
        return f'<script>window.close()</script>Erro: {request.args.get("error_description", error)}', 400

    code  = request.args.get('code')
    state = request.args.get('state')
    if state != session.get('oauth_state'):
        return '<script>window.close()</script>Estado inválido.', 400

    # Exchange code for tokens
    token_data = urllib.parse.urlencode({
        'client_id':     AZURE_CLIENT_ID,
        'client_secret': AZURE_CLIENT_SECRET,
        'code':          code,
        'redirect_uri':  APP_BASE_URL + '/auth/callback',
        'grant_type':    'authorization_code',
        'scope':         SP_SCOPES
    }).encode()
    req = urllib.request.Request(
        f'https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/token',
        data=token_data, method='POST')
    req.add_header('Content-Type', 'application/x-www-form-urlencoded')
    with urllib.request.urlopen(req) as resp:
        tokens = json.loads(resp.read())

    access_token  = tokens.get('access_token', '')
    refresh_token = tokens.get('refresh_token', '')

    # Store in session AND in short-lived memory token for popup→parent handoff
    session['sp_access_token']  = access_token
    session['sp_refresh_token'] = refresh_token

    # Generate a short token the popup passes to the parent via postMessage
    short = secrets.token_urlsafe(8)
    _TOKEN_STORE[short] = access_token

    # Popup page: sends token to parent window and closes
    return f'''<!DOCTYPE html>
<html><head><title>Ligado</title></head>
<body style="font-family:sans-serif;display:flex;align-items:center;justify-content:center;
  height:100vh;margin:0;background:#f0ede6">
<div style="text-align:center;background:#fff;padding:2rem 2.5rem;border-radius:14px;
  box-shadow:0 4px 20px rgba(0,0,0,.1)">
  <div style="font-size:2.5rem;margin-bottom:.75rem">✅</div>
  <div style="font-size:15px;font-weight:600;color:#1e3a28">Conta Microsoft ligada!</div>
  <div style="font-size:12px;color:#6b8f74;margin-top:6px">A fechar...</div>
</div>
<script>
  // Send token to parent window, then close popup
  if (window.opener) {{
    window.opener.postMessage({{type:'SP_AUTH_OK', token:'{short}'}}, '*');
  }}
  setTimeout(() => window.close(), 1000);
</script>
</body></html>'''

@app.route('/auth/sp-token', methods=['POST'])
def sp_token_exchange():
    """Parent window calls this to exchange short token for access token stored in session."""
    short = request.get_json(force=True).get('token','')
    access_token = _TOKEN_STORE.pop(short, None)  # one-time use
    if not access_token:
        return jsonify({'ok': False, 'error': 'token inválido ou expirado'}), 400
    session['sp_access_token'] = access_token
    return jsonify({'ok': True})

@app.route('/auth/sp-status')
def sp_status():
    if not authed(): return jsonify({'connected': False})
    return jsonify({'connected': bool(session.get('sp_access_token'))})

@app.route('/auth/sp-logout')
def sp_logout():
    session.pop('sp_access_token', None)
    session.pop('sp_refresh_token', None)
    return jsonify({'ok': True})

def graph_get(path):
    token = session.get('sp_access_token','')
    # Replace any control characters in the URL
    url = 'https://graph.microsoft.com/v1.0' + path
    url = url.encode('ascii', 'ignore').decode('ascii')
    req = urllib.request.Request(url,
        headers={'Authorization': 'Bearer ' + token, 'Accept': 'application/json'})
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())

@app.route('/api/sp-sites')
def sp_sites():
    if not authed() or not session.get('sp_access_token'): return jsonify({'error':'not_authenticated'}),401
    try:
        data = graph_get('/sites?search=*')
        sites = [{'id':s['id'],'name':s.get('displayName',''),'url':s.get('webUrl','')}
                 for s in data.get('value',[])]
        return jsonify({'sites': sites})
    except Exception as e: return jsonify({'error':str(e)}),500

@app.route('/api/sp-drives')
def sp_drives():
    if not authed() or not session.get('sp_access_token'): return jsonify({'error':'not_authenticated'}),401
    site_id = (request.args.get('site_id') or '').strip()
    try:
        data = graph_get(f'/sites/{urllib.parse.quote(site_id, safe=",")}/drives')
        drives = [{'id':d['id'],'name':d.get('name','')} for d in data.get('value',[])]
        return jsonify({'drives': drives})
    except Exception as e: return jsonify({'error':str(e)}),500

@app.route('/api/sp-folders')
def sp_folders():
    if not authed() or not session.get('sp_access_token'): return jsonify({'error':'not_authenticated'}),401
    drive_id  = request.args.get('drive_id','')
    folder_id = request.args.get('folder_id','root')
    # Clean IDs — remove any control characters or whitespace
    drive_id  = drive_id.strip()
    folder_id = folder_id.strip()
    try:
        # Use simpler query without $filter to avoid encoding issues
        path = f'/drives/{urllib.parse.quote(drive_id, safe="")}/items/{urllib.parse.quote(folder_id, safe="")}/children?$select=id,name,folder,file'
        data = graph_get(path)
        folders = [{'id':i['id'],'name':i['name']} for i in data.get('value',[]) if 'folder' in i]
        return jsonify({'folders':folders,'parent_id':folder_id})
    except Exception as e: return jsonify({'error':str(e)}),500

@app.route('/api/sp-upload', methods=['POST'])
def sp_upload():
    if not authed() or not session.get('sp_access_token'): return jsonify({'error':'not_authenticated'}),401
    try:
        data      = request.get_json(force=True)
        filename  = data.get('filename','orcamento.xlsx')
        drive_id  = data.get('drive_id')
        folder_id = data.get('folder_id','root')
        if not drive_id: return jsonify({'error':'Selecciona uma biblioteca.'}),400

        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(data['file_base64'])))
        for it in data.get('prices',[]):
            if it['sheet'] not in wb.sheetnames: continue
            ws = wb[it['sheet']]; r = it['row']
            if it['priceCol']>=0: ws.cell(row=r,column=it['priceCol']+1).value=it['price']
            if it.get('total') is not None and it['totalCol']>=0:
                ws.cell(row=r,column=it['totalCol']+1).value=it['total']
        out = io.BytesIO(); wb.save(out); excel_bytes = out.getvalue()

        out_name = filename.rsplit('.',1)[0]+'_preenchido.xlsx'
        token = session['sp_access_token']
        upload_url = f'https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{folder_id}:/{out_name}:/content'
        req = urllib.request.Request(upload_url, data=excel_bytes, method='PUT')
        req.add_header('Authorization','Bearer '+token)
        req.add_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        with urllib.request.urlopen(req) as resp:
            result = json.loads(resp.read())
        return jsonify({'ok':True,'name':result.get('name'),'url':result.get('webUrl')})
    except urllib.error.HTTPError as e:
        return jsonify({'error':f'Graph API {e.code}: {e.read().decode()[:300]}'}),500
    except Exception as e:
        return jsonify({'error':str(e)}),500

@app.route('/api/fill-excel', methods=['POST'])
def fill_excel():
    if not authed(): return 'Nao autorizado',401
    try:
        data=request.get_json(force=True); filename=data.get('filename','orcamento.xlsx')
        wb=openpyxl.load_workbook(io.BytesIO(base64.b64decode(data['file_base64'])))
        for it in data.get('prices',[]):
            if it['sheet'] not in wb.sheetnames: continue
            ws=wb[it['sheet']]; r=it['row']
            if it['priceCol']>=0: ws.cell(row=r,column=it['priceCol']+1).value=it['price']
            if it.get('total') is not None and it['totalCol']>=0:
                ws.cell(row=r,column=it['totalCol']+1).value=it['total']
        out=io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,download_name=filename.rsplit('.',1)[0]+'_preenchido.xlsx')
    except Exception as e: return str(e),500

@app.route('/debug')
def debug():
    return json.dumps({'BASE_DIR':BASE_DIR,'STATIC_DIR':STATIC_DIR,
        'base_files':os.listdir(BASE_DIR),
        'static_exists':os.path.exists(os.path.join(BASE_DIR,'static')),
        'azure_configured':bool(AZURE_CLIENT_ID)},indent=2),200,{'Content-Type':'application/json'}

if __name__=='__main__':
    app.run(host='0.0.0.0',port=int(os.environ.get('PORT',5000)))
